import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles import Font


def remove_blank_values(list_of_values):

    new_list = []
    for value in list_of_values:
        value = value.strip()
        if value !='.' and value!='':
            # print(value)
            new_list.append(value)
        
    return new_list

def change_to_accounting_format(excel_file, output_file):
    """
    This function reads an Excel file and if there is a ',' or '.' in numbers so it changes them to accounting format.

    Args:
        excel_file (str): The path to the Excel file.

    Returns:
        None.
    """

    wb = openpyxl.load_workbook(excel_file)
    for sheet in wb.worksheets:
        for row in range(1, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row, col)
                if isinstance(cell.value, str):
                    if ',' in cell.value or '.' in cell.value:
                        try:
                            cell.value = float(cell.value.replace(',', ''))
                            cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                        except ValueError:
                            pass
                    

    wb.save(output_file+'.xlsx')


def text_to_excel(filename,output_file):
    with open(filename, 'r') as file:
        lines = file.readlines()

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for j, line in enumerate(lines):
        

        line = line.strip()

        columns = []
        current_column = ''
        for char in line:
            current_column += char
            if current_column != ' ':
                if '  ' in current_column:
                    columns.append(current_column[:-2])
                    current_column = ''

        if current_column.strip() != '':
            columns.append(current_column.strip())
        new_columns = remove_blank_values(columns)
        try:
            # print(columns)
            worksheet.append(new_columns)
            
        except Exception as e:
            pass

        

    workbook.save(f'{output_file}.xlsx')
    return workbook

if __name__ == '__main__':
    print("""

#   _   _  ___ _   ___ __  __ __ ___                                                                    
#  | | | || __| | / _//__\|  V  | __|                                                                   
#  | 'V' || _|| || \_| \/ | \_/ | _|                                                                    
#  !_/_\_!|___|___\__/\__/|_|_|_|___|   _____   ________ _      __ ___ __  _ ___ ___  __ _____ __  ___  
#  |_   _| __\ \_/ /_   _| |_   _/__\  | __\ \_/ / _/ __| |    / _] __|  \| | __| _ \/  \_   _/__\| _ \ 
#    | | | _| > , <  | |     | || \/ | | _| > , < \_| _|| |_  | [/\ _|| | ' | _|| v / /\ || || \/ | v / 
#   _|_|_|___/_/_\_\_|_| __  |_| \__/  |___/_/_\_\__/___|___| _\__/___|_|\__|___|_|_\_||_||_| \__/|_|_\ 
#  |  V  |/__\| _ \_   _/  \|_  |/  \  | || |/  \ /' _//' _/ /  \|  \| | |                              
#  | \_/ | \/ | v / | || /\ |/ /| /\ | | >< | /\ |`._`.`._`.| /\ | | ' | |                              
#  |_| |_|\__/|_|_\ |_||_||_|___|_||_| |_||_|_||_||___/|___/|_||_|_|\__|_|                              
                                                                                                                                                                                                  
""")
    
    print('\t\t\tGithub: https://github.com/MortazaHassani')
    print(("_" * 40))
    user_file = input('Enter input text file name [enter for default VENPAY]: ')
    if isinstance(user_file, str) or user_file == '\n':
        user_file = 'VENPAY'
    
    output_file = input('Enter output file name [enter for default VENPAY]: ')
    if isinstance(output_file, str) or output_file == '\n':
        output_file = 'VENPAY'
    
    text_to_excel(user_file+'.TXT', output_file)
    change_to_accounting_format(f'{user_file}.xlsx',output_file)
    print('\t\t\t*** Success ***')
    zzz = input('Press any key to Exit!')
